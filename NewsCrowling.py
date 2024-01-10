import os
import sys
import urllib.request
import json
import re
import pandas as pd
from datetime import datetime, timedelta

today = datetime.now()
strDate = today.strftime('%Y.%m.%d')

def koreng(text):
    text_chg = text
    sign = ['&quot;', '&apos;', '&amp;', '<b>', '</b>', '  ']
    for i in sign:
        text_chg = text_chg.replace(i,'')
    return text_chg


def text_to_date(text):
    text_2 = text.replace(' +0900','')
    date = datetime.strptime(str(text_2), "%a, %d %b %Y %H:%M:%S")
    return date


def searching(client_id, client_secret, file_path):
    client_id = client_id
    client_secret = client_secret
    file_path = file_path
    df_company = pd.read_excel(file_path)

    com_list=[]
    key_list=[]
    tlist = []
    llist = []
    dlist = []
    plist = []
    total_num = 0

    for keyword in df_company.values:
        print(keyword)
        word = keyword[1]

        for pagenum in range(1,1000,100):
#         for pagenum in range(1,200,100):
            try:
                encText = urllib.parse.quote(word)
                url = "https://openapi.naver.com/v1/search/news?query=" + encText + "&display=100&sort=date&start="+str(pagenum)
#                 print(url)

                request = urllib.request.Request(url)
                request.add_header("X-Naver-Client-Id",client_id)
                request.add_header("X-Naver-Client-Secret",client_secret)
                response = urllib.request.urlopen(request)
                rescode = response.getcode()

                if(rescode==200):
                    response_body = response.read()
                    jtemp = response_body.decode('utf-8')
                    jdata= json.loads(jtemp)
                    if total_num != int(jdata['total']):
                        total_num = int(jdata['total'])
                        print(total_num)

                    for temp in jdata['items']:
                        tdata = str(temp['title'])
                        tdata = koreng(tdata)
                        ldata = temp['originallink']
                        pdata = temp['pubDate']
                        ddata = temp['description']
                        ddata = koreng(ddata)
                        tdata = tdata.replace(',',' ')
                        pdata = text_to_date(pdata)

                        com_list.append(keyword[0])
                        key_list.append(keyword[1])
                        tlist.append(tdata)
                        llist.append(ldata)
                        dlist.append(ddata)
                        plist.append(pdata)
                else:
                    print("Error Code:" + rescode)

            except:
                print('Error')


    result =[]
    for temp in range(len(tlist)):
        temp1 = []
        temp1.append(com_list[temp])
        temp1.append(key_list[temp])
        temp1.append(tlist[temp])
        temp1.append(dlist[temp])
        temp1.append(llist[temp])
        temp1.append(plist[temp])

        result.append(temp1)

    return result

def crowling(path, date):
    import account_info
    acc = account_info.Account()
    client_id = acc.id_info()
    client_secret = acc.sec_info()
    file_path = path
    news_list = searching(client_id, client_secret, file_path)
    df_result = pd.DataFrame(news_list, columns=['회사명','키워드','뉴스제목','내용','링크주소','날짜'])

    date = date
    search_date = date + ' 09:00:00'
    search_date = datetime.strptime(search_date, "%Y.%m.%d %H:%M:%S")
    date = date.replace('.','')
    today = datetime.now()
    today = today.strftime('%Y%m%d')

    df_result_2 = df_result[df_result['날짜'] >= search_date]
    df_result_2.to_excel('./{0}_{1}_News.xlsx'.format(date, today), index=False)
    
    import openpyxl
    wb = openpyxl.load_workbook('./{0}_{1}_News.xlsx'.format(date, today))
    ws = wb.active

    for cell in ws["E"]:
        if cell == ws['E1']:
            pass
        else:
            cell.value = cell.value
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    wb.save('./{0}_{1}_News.xlsx'.format(date, today))
    print(df_result_2.info())
    
    
#----------------------------------------------
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import *
from PyQt5 import uic

from datetime import datetime, timedelta


class MyWindow(QMainWindow, QWidget):    
    def __init__(self):
        super().__init__()
        self.setupUI()

    def setupUI(self):
        self.setGeometry(100, 100, 420, 600)
        self.setWindowTitle("뉴스크롤링")
        
        # 파일 불러오기
        sub_1 = QLabel("<키워드 파일을 선택해주세요>",self)
        sub_1.setGeometry(30, 10, 370, 20)
        self.text_1 = QTextEdit(self)
        self.text_1.setGeometry(20, 30, 370, 50)
        self.text_1.setFont(QtGui.QFont("",10))
        self.text_1.setText("선택한 파일 경로가 여기에 표시됩니다.")
        self.btn_1 = QPushButton("불러오기",self)
        self.btn_1.setGeometry(QtCore.QRect(320, 80, 75, 25))
        self.btn_1.clicked.connect(self.btn_fun_FileLoad)
        
        # 날짜 정하기
        sub_2 = QLabel("<시작 날짜를 선택하세요>",self)
        sub_2.setGeometry(30, 140, 370, 20)
        self.date = QLabel("Today",self)
        self.date.setGeometry(300, 140, 100, 20)
        self.date.setFont(QtGui.QFont("", 10, QtGui.QFont.Bold))
        self.date.setStyleSheet("Color : blue")
        self.dateEdit = QCalendarWidget(self)
        self.dateEdit.setGeometry(20, 160, 370, 200)
        self.dateEdit.setGridVisible(True)
        self.dateEdit.selectionChanged.connect(self.calendar_change)

        # 크롤링 실행
        sub_3 = QLabel("<프로그램을 실행시 선택하세요>",self)
        sub_3.setGeometry(30, 400, 370, 20)

        self.progressBar = QProgressBar(self)
        self.progressBar.setGeometry(40, 420, 360, 25)
        self.progressBar.setProperty("value", 0)
        self.progressBar.setTextVisible(True)
        self.progressBar.setObjectName("progressBar")
        
        self.btn_2 = QPushButton("실행", self)
        self.btn_2.setGeometry(30, 450, 370, 30)
        self.btn_2.clicked.connect(self.btn_action)
        self.btn_2.setObjectName("pushButton_2")

        # 프로그램 종료
        sub_4 = QLabel("<프로그램 종료 시 선택하세요>",self)
        sub_4.setGeometry(30, 510, 370, 20)
        self.btn_3 = QPushButton("프로그램 종료",self)
        self.btn_3.setGeometry(30, 530, 370, 30)
        self.btn_3.clicked.connect(self.close)

        
    # 버튼 실행 관련 코드
    def btn_fun_FileLoad(self):
        filePath = QFileDialog.getOpenFileName(self,"File Load", "", 'Excel(*xls *xlsx);; CSV(*csv);; All File(*)')
        if filePath[0]:
            global fileadd
            fileadd = filePath[0]
            self.text_1.setText(f"{filePath[0]}")
            try:
                df_company = pd.read_excel('{0}'.format(filePath[0]))
            except:
                self.text_1.setText("파일을 다시 선택하세요")


    def calendar_change(self):
        global strDate
        cal_date = self.dateEdit.selectedDate()
        strDate = cal_date.toString('yyyy.MM.dd')
        self.date.setText(strDate)
    
    def btn_action(self):
        self.progressBar.setProperty("value", 0)
        crowling(fileadd, strDate)
        self.progressBar.setProperty("value", 100)
        self.progressBar.setTextVisible(True)
            
if __name__ == "__main__":
    import sys
    import pandas as pd
    app = QApplication(sys.argv)
    mywindow = MyWindow()
    mywindow.show()
    app.exec_()
    print(fileadd, strDate)    